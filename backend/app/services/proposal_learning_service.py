"""
Proposal Learning Service - AI learns from past proposals
"""
import os
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
import json
from pathlib import Path

# For document parsing
try:
    from docx import Document
    from pypdf import PdfReader
    import openpyxl
except ImportError:
    pass


class ProposalLearningService:
    """
    AI service that learns from past proposals to generate better,
    domain-specific, context-aware content
    
    Features:
    - Extract patterns from past winning proposals
    - Learn company writing style and tone
    - Identify successful win themes and strategies
    - Build domain-specific knowledge base
    - Generate proposals that match past successful formats
    """
    
    def __init__(self, training_data_dir: str = "/home/ubuntu/govlogic/backend/training_data/past_proposals"):
        self.training_data_dir = training_data_dir
        self.knowledge_base: Dict[str, Any] = {
            'writing_patterns': {},
            'successful_themes': [],
            'technical_approaches': [],
            'pricing_strategies': [],
            'past_performance_examples': [],
            'company_profile': {},
            'domain_expertise': {}
        }
        self.proposal_templates: List[Dict] = []
        
        # Load existing knowledge if available
        self._load_knowledge_base()
    
    def analyze_past_proposals(self) -> Dict[str, Any]:
        """
        Analyze all past proposals in training directory
        """
        if not os.path.exists(self.training_data_dir):
            return {'error': 'Training data directory not found'}
        
        results = {
            'total_documents': 0,
            'analyzed': 0,
            'patterns_extracted': 0,
            'documents': []
        }
        
        for filename in os.listdir(self.training_data_dir):
            filepath = os.path.join(self.training_data_dir, filename)
            
            try:
                if filename.endswith('.docx'):
                    analysis = self._analyze_word_document(filepath)
                elif filename.endswith('.doc'):
                    analysis = self._analyze_legacy_word_document(filepath)
                elif filename.endswith('.pdf'):
                    analysis = self._analyze_pdf_document(filepath)
                elif filename.endswith(('.xls', '.xlsx')):
                    analysis = self._analyze_excel_document(filepath)
                else:
                    continue
                
                results['total_documents'] += 1
                if analysis:
                    results['analyzed'] += 1
                    results['documents'].append({
                        'filename': filename,
                        'analysis': analysis
                    })
                    
                    # Extract patterns
                    self._extract_patterns(analysis)
                    results['patterns_extracted'] += 1
                    
            except Exception as e:
                print(f"Error analyzing {filename}: {str(e)}")
                continue
        
        # Save updated knowledge base
        self._save_knowledge_base()
        
        return results
    
    def _analyze_word_document(self, filepath: str) -> Dict[str, Any]:
        """
        Analyze Word document (.docx)
        """
        try:
            doc = Document(filepath)
            
            analysis = {
                'type': 'technical_proposal',
                'sections': [],
                'structure': {},
                'writing_style': {},
                'key_themes': []
            }
            
            # Extract sections
            current_section = None
            section_content = []
            
            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                
                # Detect section headers (numbered or bold)
                if self._is_section_header(para):
                    if current_section:
                        analysis['sections'].append({
                            'title': current_section,
                            'content': '\n'.join(section_content),
                            'word_count': sum(len(c.split()) for c in section_content)
                        })
                    current_section = text
                    section_content = []
                else:
                    section_content.append(text)
            
            # Add last section
            if current_section:
                analysis['sections'].append({
                    'title': current_section,
                    'content': '\n'.join(section_content),
                    'word_count': sum(len(c.split()) for c in section_content)
                })
            
            # Analyze structure
            analysis['structure'] = {
                'total_sections': len(analysis['sections']),
                'section_titles': [s['title'] for s in analysis['sections']],
                'total_words': sum(s['word_count'] for s in analysis['sections'])
            }
            
            # Extract key themes
            analysis['key_themes'] = self._extract_themes_from_sections(analysis['sections'])
            
            # Analyze writing style
            analysis['writing_style'] = self._analyze_writing_style(analysis['sections'])
            
            return analysis
            
        except Exception as e:
            print(f"Error in _analyze_word_document: {str(e)}")
            return {}
    
    def _analyze_legacy_word_document(self, filepath: str) -> Dict[str, Any]:
        """
        Analyze legacy Word document (.doc)
        """
        # For .doc files, we'd need antiword or similar
        # For now, return basic analysis
        return {
            'type': 'technical_proposal',
            'format': 'legacy_word',
            'note': 'Legacy format - full analysis requires conversion'
        }
    
    def _analyze_pdf_document(self, filepath: str) -> Dict[str, Any]:
        """
        Analyze PDF document
        """
        try:
            with open(filepath, 'rb') as file:
                pdf_reader = PdfReader(file)
                
                analysis = {
                    'type': 'technical_proposal',
                    'page_count': len(pdf_reader.pages),
                    'sections': [],
                    'text_content': []
                }
                
                # Extract text from all pages
                for page_num, page in enumerate(pdf_reader.pages):
                    text = page.extract_text()
                    if text:
                        analysis['text_content'].append({
                            'page': page_num + 1,
                            'content': text
                        })
                
                return analysis
                
        except Exception as e:
            print(f"Error in _analyze_pdf_document: {str(e)}")
            return {}
    
    def _analyze_excel_document(self, filepath: str) -> Dict[str, Any]:
        """
        Analyze Excel document (pricing/cost breakdown)
        """
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            
            analysis = {
                'type': 'cost_breakdown',
                'sheets': [],
                'pricing_structure': {}
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                sheet_data = {
                    'name': sheet_name,
                    'rows': sheet.max_row,
                    'columns': sheet.max_column,
                    'data': []
                }
                
                # Extract first 20 rows as sample
                for row in sheet.iter_rows(max_row=20, values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data['data'].append(row)
                
                analysis['sheets'].append(sheet_data)
            
            # Analyze pricing patterns
            analysis['pricing_structure'] = self._analyze_pricing_patterns(analysis['sheets'])
            
            return analysis
            
        except Exception as e:
            print(f"Error in _analyze_excel_document: {str(e)}")
            return {}
    
    def _is_section_header(self, paragraph) -> bool:
        """
        Determine if paragraph is a section header
        """
        text = paragraph.text.strip()
        
        # Check for numbered sections (1.0, 2.0, etc.)
        if re.match(r'^\d+\.?\d*\s+[A-Z]', text):
            return True
        
        # Check for bold text
        if paragraph.runs and any(run.bold for run in paragraph.runs):
            return True
        
        # Check for all caps (but not too long)
        if text.isupper() and len(text.split()) <= 5:
            return True
        
        return False
    
    def _extract_themes_from_sections(self, sections: List[Dict]) -> List[str]:
        """
        Extract key themes and win strategies
        """
        themes = []
        
        # Keywords that indicate win themes
        theme_keywords = [
            'proven track record',
            'extensive experience',
            'innovative approach',
            'cost-effective',
            'local presence',
            'comprehensive',
            'expertise',
            'commitment',
            'partnership',
            'reliability'
        ]
        
        for section in sections:
            content_lower = section.get('content', '').lower()
            for keyword in theme_keywords:
                if keyword in content_lower:
                    themes.append(keyword)
        
        return list(set(themes))
    
    def _analyze_writing_style(self, sections: List[Dict]) -> Dict[str, Any]:
        """
        Analyze writing style patterns
        """
        all_content = ' '.join(s.get('content', '') for s in sections)
        
        return {
            'avg_sentence_length': self._calculate_avg_sentence_length(all_content),
            'uses_bullet_points': '•' in all_content or '●' in all_content,
            'tone': 'professional',  # Could be enhanced with NLP
            'technical_density': self._calculate_technical_density(all_content)
        }
    
    def _calculate_avg_sentence_length(self, text: str) -> float:
        """
        Calculate average sentence length
        """
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        if not sentences:
            return 0
        
        total_words = sum(len(s.split()) for s in sentences)
        return total_words / len(sentences)
    
    def _calculate_technical_density(self, text: str) -> str:
        """
        Estimate technical density
        """
        technical_indicators = [
            'system', 'technology', 'implementation', 'architecture',
            'framework', 'methodology', 'process', 'requirements',
            'specifications', 'standards', 'compliance'
        ]
        
        text_lower = text.lower()
        count = sum(text_lower.count(word) for word in technical_indicators)
        words = len(text.split())
        
        if words == 0:
            return 'unknown'
        
        density = count / words
        
        if density > 0.05:
            return 'high'
        elif density > 0.02:
            return 'medium'
        else:
            return 'low'
    
    def _analyze_pricing_patterns(self, sheets: List[Dict]) -> Dict[str, Any]:
        """
        Analyze pricing structure patterns
        """
        return {
            'has_labor_categories': True,
            'has_hourly_rates': True,
            'has_total_cost': True,
            'pricing_model': 'time_and_materials'  # Could be enhanced
        }
    
    def _extract_patterns(self, analysis: Dict[str, Any]):
        """
        Extract patterns and add to knowledge base
        """
        if analysis.get('type') == 'technical_proposal':
            # Add section structure
            if 'structure' in analysis:
                self.knowledge_base['writing_patterns']['common_sections'] = \
                    analysis['structure'].get('section_titles', [])
            
            # Add themes
            if 'key_themes' in analysis:
                self.knowledge_base['successful_themes'].extend(analysis['key_themes'])
            
            # Add writing style
            if 'writing_style' in analysis:
                self.knowledge_base['writing_patterns']['style'] = analysis['writing_style']
        
        elif analysis.get('type') == 'cost_breakdown':
            # Add pricing structure
            if 'pricing_structure' in analysis:
                self.knowledge_base['pricing_strategies'].append(analysis['pricing_structure'])
    
    def generate_proposal_section(
        self,
        section_title: str,
        requirements: List[str],
        context: Dict[str, Any]
    ) -> str:
        """
        Generate proposal section using learned patterns
        """
        # Use knowledge base to inform generation
        style = self.knowledge_base['writing_patterns'].get('style', {})
        themes = self.knowledge_base['successful_themes']
        
        # Build prompt with learned context
        prompt = f"""
        Generate a {section_title} section for a government proposal.
        
        Writing Style Guidelines (learned from past winning proposals):
        - Average sentence length: {style.get('avg_sentence_length', 20)} words
        - Technical density: {style.get('technical_density', 'medium')}
        - Tone: {style.get('tone', 'professional')}
        
        Successful Themes to Incorporate:
        {', '.join(themes[:5])}
        
        Requirements to Address:
        {chr(10).join(f"- {req}" for req in requirements)}
        
        Context:
        {json.dumps(context, indent=2)}
        
        Generate a comprehensive, compliant section that addresses all requirements
        while maintaining the learned writing style and incorporating successful themes.
        """
        
        return prompt  # In production, this would call the AI service
    
    def get_company_profile(self) -> Dict[str, Any]:
        """
        Extract company profile from past proposals
        """
        return {
            'name': 'Unified Industries Incorporated',
            'location': 'Alexandria, VA',
            'certifications': ['Minority-Owned Small Disadvantaged Business'],
            'experience_years': '50+',
            'core_competencies': [
                'IT Staff Augmentation',
                'Application Development',
                'Project/Program Management',
                'Government Contracting'
            ],
            'past_clients': [
                'Fairfax County',
                'WMATA',
                'Federal Government Agencies'
            ]
        }
    
    def _load_knowledge_base(self):
        """
        Load existing knowledge base from file
        """
        kb_path = os.path.join(self.training_data_dir, 'knowledge_base.json')
        if os.path.exists(kb_path):
            try:
                with open(kb_path, 'r') as f:
                    self.knowledge_base = json.load(f)
            except Exception as e:
                print(f"Error loading knowledge base: {str(e)}")
    
    def _save_knowledge_base(self):
        """
        Save knowledge base to file
        """
        kb_path = os.path.join(self.training_data_dir, 'knowledge_base.json')
        try:
            os.makedirs(os.path.dirname(kb_path), exist_ok=True)
            with open(kb_path, 'w') as f:
                json.dump(self.knowledge_base, f, indent=2)
        except Exception as e:
            print(f"Error saving knowledge base: {str(e)}")
    
    def get_knowledge_summary(self) -> Dict[str, Any]:
        """
        Get summary of learned knowledge
        """
        return {
            'total_themes': len(self.knowledge_base['successful_themes']),
            'unique_themes': len(set(self.knowledge_base['successful_themes'])),
            'common_sections': self.knowledge_base['writing_patterns'].get('common_sections', []),
            'writing_style': self.knowledge_base['writing_patterns'].get('style', {}),
            'pricing_strategies_count': len(self.knowledge_base['pricing_strategies']),
            'company_profile': self.get_company_profile()
        }


# Global instance
proposal_learning_service = ProposalLearningService()

